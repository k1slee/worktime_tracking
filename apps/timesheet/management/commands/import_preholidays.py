from django.core.management.base import BaseCommand
from openpyxl import load_workbook
from datetime import datetime, date
from apps.timesheet.models import Holiday

class Command(BaseCommand):
    help = "Импорт предпраздничных дней из Excel"

    def add_arguments(self, parser):
        parser.add_argument("filename", type=str, help="Путь к Excel файлу")

    def handle(self, *args, **options):
        filename = options["filename"]
        wb = load_workbook(filename)
        sheet = wb.active

        # начиная с A2 до J... и вниз
        for row in sheet.iter_rows(min_row=2, max_col=10):
            for cell in row:
                if cell.value:
                    if isinstance(cell.value, (datetime, date)):
                        day = cell.value.date() if isinstance(cell.value, datetime) else cell.value
                        Holiday.objects.get_or_create(
                            date=day,
                            defaults={"type": "preholiday", "name": "Предпраздничный день"}
                        )
                    else:
                        self.stdout.write(self.style.WARNING(f"Пропущено: {cell.value}"))

        self.stdout.write(self.style.SUCCESS("Импорт предпраздничных дней завершён"))
